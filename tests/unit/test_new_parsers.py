"""Unit tests for the new document parsers (CSV, Excel, JSON)."""

import io
import openpyxl
import pytest

from app.analysis.parsers import parser_factory
from app.core.errors import UnparsableDocumentError


def test_csv_parser() -> None:
    csv_bytes = b"Name,Skill\nJohn Doe,Python\n"
    text = parser_factory.parse("resume.csv", csv_bytes)
    assert "John Doe" in text
    assert "Python" in text


def test_json_parser() -> None:
    json_bytes = b'{"name": "John Doe", "skills": ["Python"]}'
    text = parser_factory.parse("resume.json", json_bytes)
    assert "John Doe" in text
    assert "Python" in text


def test_excel_parser() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "John Doe"
    ws['B1'] = "Python"
    out = io.BytesIO()
    wb.save(out)
    excel_bytes = out.getvalue()

    text = parser_factory.parse("resume.xlsx", excel_bytes)
    assert "John Doe" in text
    assert "Python" in text
